import pages.sourcing_table.dataframe as dataframe
import dash
from dash import html
from app import app
import dash_ag_grid as dag
from dash.dependencies import Input, Output, State
import pandas as pd
from dash.exceptions import PreventUpdate
from dash import dcc
import dash_bootstrap_components as dbc
import utils
import os
import pages.archive.archive as archive
import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import numbers


service_name=utils.service_name
service_code=utils.service_code
role=utils.role
region=utils.region
info=utils.info



id=utils.id_factory('archive_simulation_ope')

def field_quarter(year,quarter):
    return f"{year} Q{quarter}"

def getColDefSimu(year,next_year,next_next_year):
    #Transform the year from the format YYYY to the format YY
    year_yy= year % 100
    next_year_yy= next_year % 100
    next_next_year_yy=next_next_year % 100

    #Variable used to name the different columns for the current year
    Q1 = f"1Q {year_yy}"
    Q2 = f"2Q {year_yy}"
    Q3 = f"3Q {year_yy}"
    Q4 = f"4Q {year_yy}"

    #Variable used to name the different columns for the next year
    Q1_next = f"1Q {next_year_yy}"
    Q2_next = f"2Q {next_year_yy}"
    Q3_next = f"3Q {next_year_yy}"
    Q4_next = f"4Q {next_year_yy}"

    #Variable used to name the different columns for the next next year
    Q1_two = f"1Q {next_next_year_yy}"
    Q2_two = f"2Q {next_next_year_yy}"
    Q3_two = f"3Q {next_next_year_yy}"
    Q4_two = f"4Q {next_next_year_yy}"
        
    col_def_simu=[
                {
                "headerName": "",
                "children":[
                    {
                        "headerName":service_code,
                        "field": service_code,                  
                        "width":115,
                        "editable": False,
                        "cellClassRules": {"spanned-row": "params.value"},
                        "cellStyle":{"styleConditions":[{"condition": f"params.data.{info} == '110A TOPs' || params.data.{info} == 'AD10 TOPs'", "style":{"color":"rgba(0,0,0,0)"}}],"defaultStyle": {"color": "black"}},
                        "hide":True,
                    },
                    {
                        "headerName":service_name,
                        "field":service_name,  
                        "width":105,
                        'floatingFilter': False,
                        "filter":True,
                        "filterParams":{"maxNumConditions":1,'buttons': ['reset']},
                        "editable": False,
                        "cellClassRules": {"spanned-row": "params.value"},
                        "cellStyle":{"styleConditions":[{"condition": f"params.data.{info} == '110A TOPs' || params.data.{info} == 'AD10 TOPs'", "style":{"color":"rgb(0,0,0,0)"}}],"defaultStyle": {"color": "black"}},
                        "hide":True,
                        
                        
                    },
                    {
                        "headerName":role,
                        "field":role,     
                        "width":105,
                        'floatingFilter': False,
                        "filter":True,
                        "filterParams":{"maxNumConditions":1,'buttons': ['reset']},
                        "editable": False,
                        "cellClassRules": {"spanned-row": "params.value"},
                        "hide":True,
                        "cellStyle":{"styleConditions":[{"condition": f"params.data.{info} == '110A TOPs' || params.data.{info} == 'AD10 TOPs'", "style":{"color":"rgb(0,0,0,0)"}}],"defaultStyle": {"color": "black"}},
                        
                    },
                    {
                        "headerName":region,
                        "field":region,    
                        "width":105,
                        'floatingFilter': False,
                        "filter":True,
                        "filterParams":{"maxNumConditions":1,'buttons': ['reset']},
                        "editable": False,
                        "hide":True,
                        "cellClassRules": {"spanned-row": "params.value"},
                        "cellStyle":{"styleConditions":[{"condition": f"params.data.{info} == '110A TOPs' || params.data.{info} == 'AD10 TOPs'", "style":{"color":"rgb(0,0,0,0)"}}],"defaultStyle": {"color": "black"}},
                    },
                    {
                        "headerName":"Indicator",
                        "field":info,    
                        "width":415,
                        "editable": False,
                        "cellStyle": {'borderLeft': '3px solid black'}, 
                        'cellRendererSelector': {'function': 'cellRendererSelector(params)'},
                    },
                    
                ]
                },
                {
                    "headerName":f"{str(year)}",

                    "children":
                    [
    
                        {
                            "headerName": Q1,
                            "field": field_quarter(year,1),
                            "width":105,
                            "editable": False,  
                            "cellEditor": {"function": "NumberInput"},
                            "valueGetter": {"function" : f"valueGetter(params,{year},1)"},
                            "valueFormatter": {"function" : "valuePercentage(params)"},
                            "cellStyle": {'borderLeft': '3px solid black'}, 
                                
                        },
                        {
                            "headerName": Q2,
                            "field": field_quarter(year,2),
                            "width":105,
                            "editable": False,
                            "cellEditor": {"function": "NumberInput"},
                            "valueGetter": {"function" : f"valueGetter(params,{year},2)"},
                            "valueFormatter": {"function" : "valuePercentage(params)"},
                            "cellStyle": {'borderLeft': '1px solid lightgrey'}, 
                                                    
                            
                        },
                        {
                            "headerName": Q3,
                            "field": field_quarter(year,3),
                            "width":105,
                            "editable": False,
                            "cellEditor": {"function": "NumberInput"},
                            "valueGetter": {"function" : f"valueGetter(params,{year},3)"},
                            "valueFormatter": {"function" : "valuePercentage(params)"}, 
                            "cellStyle": {'borderLeft': '1px solid lightgrey'}, 
                            
                            
                        },
                        {
                            "headerName": Q4,
                            "field": field_quarter(year,4),
                            "width":105,
                            "editable": False,
                            "valueGetter": {"function" : f"valueGetter(params,{year},4)"},
                            "valueFormatter": {"function" : "valuePercentage(params)"},   
                            "cellStyle": {'borderLeft': '1px solid lightgrey'}, 
                            
                            
                        },
                    ]
                    },
                    {
                    "headerName":f"{str(next_year)}",

                    "children":
                    [
    
                        {
                            "headerName": Q1_next,
                            "field": field_quarter(next_year,1),
                            "width":105,
                            "editable":  False,
                            "valueGetter": {"function" : f"valueGetter(params,{next_year},1)"},
                            "valueFormatter": {"function" : "valuePercentage(params)"},
                            "cellStyle": {'borderLeft': '3px solid black'},   
                            
                            
                        },
                        {
                            "headerName": Q2_next,
                            "field": field_quarter(next_year,2),
                            "width":105,
                            "editable":  False,
                            "valueGetter": {"function" : f"valueGetter(params,{next_year},2)"},
                            "valueFormatter": {"function" : "valuePercentage(params)"},
                            "cellStyle": {'borderLeft': '1px solid lightgrey'},    
                            
                            
                        },
                        {
                            "headerName": Q3_next,
                            "field": field_quarter(next_year,3),
                            "width":105,
                            "editable":  False,
                            "valueGetter": {"function" : f"valueGetter(params,{next_year},3)"},
                            "valueFormatter": {"function" : "valuePercentage(params)"}, 
                            "cellStyle": {'borderLeft': '1px solid lightgrey'},  
                            
                            
                        },
                        {
                            "headerName": Q4_next,
                            "field": field_quarter(next_year,4),
                            "width":105,
                            "editable":  False,
                            "valueGetter": {"function" : f"valueGetter(params,{next_year},4)"},
                            "valueFormatter": {"function" : "valuePercentage(params)"},   
                            "cellStyle": {'borderLeft': '1px solid lightgrey'},                         
                        },
                    ]
                    },
                                    {
                    "headerName":f"{str(next_next_year)}",

                    "children":
                    [
    
                        {
                            "headerName": Q1_two,
                            "field": field_quarter(next_next_year,1),
                            "width":85,
                            "editable":  False,
                            "valueGetter": {"function" : f"valueGetter(params,{next_next_year},1)"},
                            "valueFormatter": {"function" : "valuePercentage(params)"}, 
                            "cellStyle": {'borderLeft': '3px solid black'},  
                            
                        },
                        {
                            "headerName": Q2_two,
                            "field": field_quarter(next_next_year,2),
                            "width":85,
                            "editable":  False,
                            "valueGetter": {"function" : f"valueGetter(params,{next_next_year},2)"},
                            "valueFormatter": {"function" : "valuePercentage(params)"},    
                            "cellStyle": {'borderLeft': '1px solid lightgrey'},
                            
                            
                        },
                        {
                            "headerName": Q3_two,
                            "field": field_quarter(next_next_year,3),
                            "width":85,
                            "editable":  False,
                            "valueGetter": {"function" : f"valueGetter(params,{next_next_year},3)"},
                            "valueFormatter": {"function" : "valuePercentage(params)"},   
                            "cellStyle": {'borderLeft': '1px solid lightgrey'},
                            
                            
                        },
                        {
                            "headerName": Q4_two,
                            "field": field_quarter(next_next_year,4),
                            "width":85,
                            "editable":  False,
                            "cellEditor": {"function": "NumberInput"},
                            "valueGetter": {"function" : f"valueGetter(params,{next_next_year},4)"},
                            "valueFormatter": {"function" : "valuePercentage(params)"},    
                            "cellStyle": {'borderLeft': '1px solid lightgrey'},                        
                        },
                    ]
                    },
                    {
                        "headerName": f"Avg {str(year_yy)}",
                        "valueGetter": {"function" : f"Math.round(((params.data['{str(year)} Q1']+params.data['{str(year)} Q2']+params.data['{str(year)} Q3']+params.data['{str(year)} Q4'])/4)*10)/10"},
                        "width":95,
                        "valueFormatter": {"function" : "valuePercentage(params)"}, 
                        "cellStyle": {'borderLeft': '3px solid black'},  
                                                                    
                    },
                    {
                        "headerName": f"Avg {str(next_year_yy)}",
                        "valueGetter": {"function" : f"Math.round(((params.data['{str(next_year)} Q1']+params.data['{str(next_year)} Q2']+params.data['{str(next_year)} Q3']+params.data['{str(next_year)} Q4'])/4)*10)/10"},
                        "width":95,
                        "valueFormatter": {"function" : "valuePercentage(params)"},    
                        "cellStyle": {'borderLeft': '1px solid lightgrey'},                                            
                    },
                    {
                        "headerName": f"Avg {str(next_next_year_yy)}",
                        "valueGetter": {"function" : f"Math.round(((params.data['{str(next_next_year)} Q1']+params.data['{str(next_next_year)} Q2']+params.data['{str(next_next_year)} Q3']+params.data['{str(next_next_year)} Q4'])/4)*10)/10"},
                        "width":95,
                        "valueFormatter": {"function" : "valuePercentage(params)"}, 
                        "cellStyle": {'borderLeft': '1px solid lightgrey'},                                               
                    },
                    {
                        "headerName": f"Var {str(year_yy)}/{str(next_year_yy)}",
                        "width":100,
                        "valueFormatter": {"function" : "valuePercentage(params)"},
                        "valueGetter": {"function" : f"Math.round(Math.round(((params.data['{str(next_year)} Q1']+params.data['{str(next_year)} Q2']+params.data['{str(next_year)} Q3']+params.data['{str(next_year)} Q4'])/4)*10)-Math.round(((params.data['{str(year)} Q1']+params.data['{str(year)} Q2']+params.data['{str(year)} Q3']+params.data['{str(year)} Q4'])/4)*10))/10"},
                        "cellStyle": {'borderLeft': '3px solid black'},
                    },
                    {
                        "headerName": f"Var {str(next_year_yy)}/{str(next_next_year_yy)}",
                        "width":100,
                        "valueFormatter": {"function" : "valuePercentage(params)"},
                        "valueGetter": {"function" : f"Math.round(Math.round(((params.data['{str(next_next_year)} Q1']+params.data['{str(next_next_year)} Q2']+params.data['{str(next_next_year)} Q3']+params.data['{str(next_next_year)} Q4'])/4)*10)-Math.round(((params.data['{str(next_year)} Q1']+params.data['{str(next_year)} Q2']+params.data['{str(next_year)} Q3']+params.data['{str(next_year)} Q4'])/4)*10))/10"},
                        "cellStyle": {'borderLeft': '1px solid lightgrey','borderRight': '2px solid black'},
                    },

            ]
    return col_def_simu


getRowStyle = {
    "styleConditions": [
        {
            "condition": f"params.data.{info} ==='In-House Workload - Total FTEs - after PRP & Approved+Not Approved'",
            "style": {"backgroundColor": "rgba(35,0,76,1)","color":"white","borderTop":"1px solid black","borderBottom":"1px solid black"},
        },
        {
            "condition": f"params.data.{info} ==='Total In-House Workforce - FTEs'",
            "style": {"backgroundColor": "rgba(35,0,76,1)","color":"white","borderTop":"1px solid black","borderBottom":"1px solid black"},
        },
        {
            "condition": f"params.data.{info} ==='In House Coverage %'",
            "style": {"backgroundColor": "rgba(35,0,76,1)","color":"white","borderTop":"1px solid black","borderBottom":"1px solid black"},
        },
        {
            "condition": f"params.data.{info} ==='110A TOPs'",
            "style": {"backgroundColor":"rgba(35,0,76,0.25)","color":"black","borderTop":"1px solid black"},
        },
        {
            "condition": f"params.data.{info} ==='AD10 TOPs'",
            "style": {"backgroundColor":"rgba(35,0,76,0.25)","color":"black","borderBottom":"1px solid black"},
        },
        {
            "condition": f"params.data.{info} ==='Total In-House Workforce - FTEs with TOPs'",
            "style": {"backgroundColor": "rgba(35,0,76,1)","color":"white","borderTop":"1px solid black","borderBottom":"1px solid black"},
        },
        {
            "condition": f"params.data.{info} ==='In House Coverage % with TOPS'",
            "style": {"backgroundColor": "rgba(35,0,76,1)","color":"white","borderTop":"1px solid black","borderBottom":"1px solid black"},
        },
        
    ],
}

defaultColDef_simu = {"resizable": True, "filter": False,"wrapHeaderText": True,"suppressMovable": True, "editable": False}

#===========================================================================================================
#                                               Definition of the layout

# Isolates existing services and links them to the corresponding code. Used in the graphical view dropdown
# We use a function rather than a variable to dynamically update the data
def serviceToCode(df):
    unique_services = df.drop_duplicates(subset=[service_name])
    service_to_code = {row[service_name]: row[service_code] for index, row in unique_services.iterrows()}
    return service_to_code



def get_Layout(year,next_year,next_next_year,data):

    
    if not isinstance(next_next_year,int):
        layout=html.Div([
            html.H2(f'Simulation (operational roles)',style={"textAlign":"center","color":"#23004C", "paddingTop":"10px"}),
            html.Hr(),
            html.P("There are currently no archives selected. Please select one using the \"Select an archive\" button.")
        ])
    
    else:
  
        col_def_simu=getColDefSimu(year,next_year,next_next_year)
        data=pd.read_json(data, orient='split') 
        data=data.round(1)       
        layout=html.Div(
            [
                html.H2('Simulation (operational roles)',style={"textAlign":"center","color":"#23004C", "paddingTop":"10px"}),
                html.Hr(),
                html.Div([
                    html.Div([
                        html.P("Service Name: ",style={'display': 'inline-block','marginRight': '5%'}),  # Ajout du texte avant le premier dropdown
                        dcc.Dropdown(
                            options=[{'label': f"{service} ({serviceToCode(data)[service]})", 'value': service} for service in data[service_name].unique()],
                            value= data[service_name].unique()[0],
                            id=id('dropdown_service_name'),
                            style={'width': '60%','display': 'inline-block','verticalAlign': 'middle'}
                        ),
                    ], style={'width': '25%', 'display': 'inline-block'}),
                    html.Div([
                        html.P("Role: ",style={'display': 'inline-block','marginRight': '5%'}),  # Ajout du texte avant le deuxième dropdown
                        dcc.Dropdown(
                            options=[{'label': r, 'value': r} for r in data[role].unique()],
                            value="ALL",
                            id=id('dropdown_role'),
                            style={'width': '60%','display': 'inline-block','verticalAlign': 'middle'}
                        ),
                    ], style={'width': '25%', 'display': 'inline-block'}),
                    html.Div([
                        html.P("Region: ",style={'display': 'inline-block','marginRight': '5%'}),  # Ajout du texte avant le troisième dropdown
                        dcc.Dropdown(
                            options=[{'label': i, 'value': i} for i in data[region].unique()],
                            value="ALL",
                            id=id('dropdown_region'),
                            style={'width': '60%','display': 'inline-block','verticalAlign': 'middle'}
                        ),
                    ], style={'width': '25%', 'display': 'inline-block'}),
                ], style={"width": "100%"}),
                dag.AgGrid(
                    id="archive_simulation_ope",
                    columnDefs=col_def_simu,
                    defaultColDef=defaultColDef_simu,
                    rowData=data.to_dict("records"),
                    dashGridOptions = {'suppressRowTransform': True,'animateRows' : False, "singleClickEdit": True,"stopEditingWhenCellsLoseFocus": True },
                    getRowStyle = getRowStyle,
                    style={'width': '100%', 'height': '775px', 'marginTop': '10px'},
                    columnSize="responsiveSizeToFit",
                    filterModel={'Service Name':{'filterType' : 'text', 'type': 'contains', 'filter': f'{data[service_name].unique()[0]}'},'Role':{'filterType' : 'text', 'type': 'contains', 'filter': 'ALL'},
                                'Region':{'filterType' : 'text', 'type': 'contains', 'filter': 'ALL'}}
                ),
                dcc.Loading(
                    id="loading_data_simu",
                    type="circle",
                    color="#7A00E6",
                    children=[               
                        html.Div(
                            [
                                dbc.Button('Download data',id=id('download_button'),n_clicks=0,class_name="me-2", style={"backgroundColor":"#7A00E6", "borderColor": "#7A00E6"}),
                                dcc.Download(id=id("download_data_func")),                     
                            ],
                        style={"textAlign":"center","paddingTop":"20px","width":"17rem", "margin":"0 auto"},
                        ),
                    ],
                ),                
            ],

        )
    return layout



#===========================================================================================================
#                                               Callbacks

@app.callback([Output(id('dropdown_role'),'options',allow_duplicate=True),
               Output(id('dropdown_role'),'value',allow_duplicate=True)],
               Input(id('dropdown_service_name'),'value'),
               State("data_archive_simu_ope","data"),
              prevent_initial_call=True)
def update_dd_role(value,data):
    """_summary_

    Args:
        value (_type_): _description_

    Returns:
        _type_: _description_
    """
    data=pd.read_json(data, orient='split')      
    if value is None:
        return data[role].unique()
    
    mask=data[service_name] == value
    simu_ope=data[mask]
    return simu_ope[role].unique(), 'ALL'


@app.callback([Output(id('dropdown_region'),'options',allow_duplicate=True),
               Output(id('dropdown_region'),'value',allow_duplicate=True)],             
              Input(id('dropdown_role'),'value'),
              State(id('dropdown_service_name'),'value'),
              State("data_archive_simu_ope","data"),
              prevent_initial_call=True)
def update_dd_region(selected_role,service,data):
    """
    _summary_

    Args:
        role (_type_): _description_
        service (_type_): _description_

    Returns:
        _type_: _description_
    """
    data=pd.read_json(data, orient='split')  
    if role is None:
        return data[region].unique()
    
    mask_role = data[role] == selected_role
    mask_service = data[service_name] == service
    filtered_df = data[mask_role & mask_service]
    if len(filtered_df[region].unique()) == 2 and filtered_df[region].unique()[0] == " ":
        return filtered_df[region].unique()," "
    return filtered_df[region].unique(),'ALL'


@app.callback(Output('archive_simulation_ope','filterModel',allow_duplicate=True),
              Input(id('dropdown_service_name'),'value'),
              prevent_initial_call=True)
def filter_service(dd_service):
    """_summary_

    Args:
        dd_service (_type_): _description_

    Returns:
        _type_: _description_
    """

    filter_role={"type": "equals", "filter" : ""}
    if dd_service is None:
        filter_name= {"type": "contains", "filter" : ""}
        return {"Service Name" : filter_name, "Role" : filter_role}
    filter_name= {"type": "equals", "filter" : dd_service}
    return {"Service Name" : filter_name, "Role" : filter_role}



@app.callback(Output('archive_simulation_ope','filterModel',allow_duplicate=True),
              Input(id('dropdown_role'),'value'),
              State(id('dropdown_service_name'),'value'),
              prevent_initial_call=True)
def filter_role(dd_role,value):
    """_summary_

    Args:
        dd_role (_type_): _description_
        value (_type_): _description_

    Returns:
        _type_: _description_
    """
    filter_name={"type": "equals", "filter" : value}
    if dd_role is None:
        filter_role= {"type": "contains", "filter" : ""}
        return {"Role" : filter_role, "Service Name": filter_name}
    
    filter_role= {"type": "equals", "filter" : dd_role}

    return {"Role" : filter_role, "Service Name": filter_name}

@app.callback(Output('archive_simulation_ope','filterModel',allow_duplicate=True),
              Input(id('dropdown_region'),'value'),
              State(id('dropdown_service_name'),'value'),
              State(id('dropdown_role'),'value'),
              prevent_initial_call=True)
def filter_region(dd_region,value_ser,value_rol):
    """_summary_

    Args:
        dd_region (_type_): _description_
        value_ser (_type_): _description_
        value_rol (_type_): _description_

    Returns:
        _type_: _description_
    """
    filter_name={"type": "equals", "filter" : value_ser}
    filter_role={"type": "equals", "filter" : value_rol}
    if dd_region is None:
        filter_region= {"type": "contains", "filter" : ""}
        return {"Role" : filter_role, "Service Name": filter_name, "Region" : filter_region}
    
    filter_region= {"type": "equals", "filter" : dd_region}

    return {"Role" : filter_role, "Service Name": filter_name, "Region": filter_region}

@app.callback(
    Output(id("download_data_func"), "data"),
    Input(id("download_button"), "n_clicks"),
    State("data_archive_simu_ope","data"),
    State("archive_date_store","data"),
    prevent_initial_call=True,
)
def func(n_clicks,data,date):
    df=pd.read_json(data, orient='split')   
    
    if date:
        value=datetime.datetime.strptime(date,"%d %B %Y")
        year=value.year
        next_year=year+1
        year_two=next_year+1

    
    # Exclure certaines valeurs selon les conditions spécifiées
    df = df[~(df[role] == "ALL")]
    df = df[~(df[region] == "ALL")]
    df = df[~(df[service_name] == "CSO")]
    df = df[~(df[service_name] == "CSO+EGDS")]
    df = df[~(df[service_name] == "ALL")]
    
    col = df.pop("Info")
    df.insert(4, "Info", col)

    list_years = [year, next_year, year_two]
    quarter_columns = [f"{y} Q{i}" for y in list_years for i in range(1, 5)]

    # Créer un fichier temporaire pour enregistrer l'Excel
    file_name = f"Report operational roles- {date}.xlsx"    
    
    df.to_excel(file_name, index=False, float_format="%.2f")

    # Charger le fichier Excel avec openpyxl
    wb = load_workbook(file_name)
    ws = wb.active

    # Appliquer le format pourcentage aux cellules des colonnes de trimestres si 'info' contient '%'
    for row_idx, info_value in enumerate(df[info], start=2):  # Démarrer à la ligne 2 pour ignorer l'en-tête
        if "%" in str(info_value):  # Vérifier si la valeur de 'info' contient "%"
            for col_name in quarter_columns:
                if col_name in df.columns:
                    col_idx = df.columns.get_loc(col_name) + 1  # Obtenir l'index de la colonne (1-based)
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:  # Appliquer le format si la cellule a une valeur
                        # Diviser la valeur par 100 pour un affichage correct
                        cell.value = cell.value / 100
                        cell.number_format = numbers.FORMAT_PERCENTAGE_00

    # Sauvegarder le fichier avec le format appliqué
    wb.save(file_name)

    # Retourner le fichier Excel formaté
    return dcc.send_file(file_name)


